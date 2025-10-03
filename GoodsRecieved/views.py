# goods_received/views.py
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Count, Sum, Avg
from django.db import transaction
from datetime import datetime, timedelta
from django.utils import timezone
import pandas as pd
from django.http import HttpResponse
import io
import logging

from .models import GoodsReceivedChina, GoodsReceivedGhana, GoodsReceivedContainer, GoodsReceivedItem
from users.models import CustomerUser
from .serializers import (
    GoodsReceivedChinaSerializer,
    GoodsReceivedGhanaSerializer,
    StatusUpdateSerializer,
    BulkStatusUpdateSerializer,
    GoodsSearchSerializer,
    GoodsStatsSerializer,
    ExcelUploadSerializer,
    BulkCreateResultSerializer,
    GoodsReceivedContainerSerializer,
    GoodsReceivedItemSerializer,
    CreateGoodsReceivedContainerSerializer,
    CreateGoodsReceivedItemSerializer,
    GoodsReceivedContainerStatsSerializer,
    MAX_SHIPPING_MARK_LENGTH
)

logger = logging.getLogger(__name__)


class BaseGoodsReceivedViewSet(viewsets.ModelViewSet):
    """
    Base ViewSet for managing goods received in warehouses.
    Lean, model-agnostic: child classes must set model_class, serializer_class and warehouse_type.
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    # generic safe filters/search/order (fields present in both models)
    filterset_fields = ['status', 'shipping_mark', 'supply_tracking']
    search_fields = ['shipping_mark', 'supply_tracking', 'description']
    ordering_fields = ['date_received', 'created_at', 'status', 'cbm', 'weight']
    ordering = ['-date_received']

    model_class = None
    serializer_class = None
    warehouse_type = None
    template_filename = None
    template_sheet_name = None

    def get_queryset(self):
        # base queryset: child classes set self.model_class and queryset property can use this
        queryset = self.model_class.objects.all()

        # date range filtering
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            try:
                date_from = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                queryset = queryset.filter(date_received__gte=date_from)
            except ValueError:
                logger.warning(f"Invalid date_from format: {date_from}")
        if date_to:
            try:
                date_to = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                queryset = queryset.filter(date_received__lte=date_to)
            except ValueError:
                logger.warning(f"Invalid date_to format: {date_to}")

        # optional search param that may include shipping_mark or supply_tracking
        search = self.request.query_params.get('q')
        if search:
            queryset = queryset.filter(
                Q(shipping_mark__icontains=search) |
                Q(supply_tracking__icontains=search) |
                Q(description__icontains=search)
            )

        return queryset

    def _apply_status_change(self, instance, new_status, reason=None):
        """
        Apply status change using the model's helper methods when available.
        Avoid referencing removed fields (notes, item_id).
        """
        # Ghana-specific transitions
        if self.warehouse_type == 'ghana':
            if new_status == 'READY_FOR_DELIVERY' and hasattr(instance, 'mark_ready_for_delivery'):
                instance.mark_ready_for_delivery()
                return
            if new_status == 'DELIVERED' and hasattr(instance, 'mark_delivered'):
                instance.mark_delivered()
                return
            # generic fallback
            instance.status = new_status
            instance.save(update_fields=['status', 'updated_at'])
            return

        # China-specific transitions (default)
        if new_status == 'READY_FOR_SHIPPING' and hasattr(instance, 'mark_ready_for_shipping'):
            instance.mark_ready_for_shipping()
            return
        if new_status == 'SHIPPED' and hasattr(instance, 'mark_shipped'):
            instance.mark_shipped()
            return
        instance.status = new_status
        instance.save(update_fields=['status', 'updated_at'])

    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """
        Update single item status.
        Uses model helper methods for business logic and atomic transaction for safety.
        """
        obj = self.get_object()
        serializer = StatusUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        new_status = serializer.validated_data['status']
        reason = serializer.validated_data.get('reason')
        try:
            self._apply_status_change(obj, new_status, reason)
            serialized = self.get_serializer(obj)
            return Response({'message': f'Status updated to {new_status}', 'item': serialized.data})
        except Exception as e:
            logger.exception("Error updating status")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], throttle_classes=[])
    def bulk_status_update(self, request):
        """
        Bulk update status for multiple supply_tracking values.
        Uses atomic transaction for safety. Throttling can be added for abuse prevention.
        """
        serializer = BulkStatusUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        item_ids = serializer.validated_data['item_ids']
        new_status = serializer.validated_data['status']
        reason = serializer.validated_data.get('reason')

        # find items by supply_tracking
        qs = self.model_class.objects.filter(supply_tracking__in=item_ids)
        updated = 0

        # simple bulk set for trivial status changes
        simple_bulkable = new_status in ['PENDING', 'CANCELLED'] and not reason
        if simple_bulkable:
            updated = qs.update(status=new_status, updated_at=timezone.now())
            return Response({'message': f'Updated {updated} items', 'updated_count': updated})

        # otherwise iterate and apply helpers
        with transaction.atomic():
            for instance in qs:
                try:
                    self._apply_status_change(instance, new_status, reason)
                    updated += 1
                except Exception as e:
                    logger.exception(f"Failed to update {instance.supply_tracking}: {e}")

        return Response({'message': f'Updated {updated} items', 'updated_count': updated})

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Warehouse statistics (aggregations)"""
        qs = self.get_queryset()

        # status names per warehouse
        if self.warehouse_type == 'ghana':
            ready_status = 'READY_FOR_DELIVERY'
            complete_status = 'DELIVERED'
        else:
            ready_status = 'READY_FOR_SHIPPING'
            complete_status = 'SHIPPED'

        agg = qs.aggregate(
            total_count=Count('id'),
            pending_count=Count('id', filter=Q(status='PENDING')),
            flagged_count=Count('id', filter=Q(status='FLAGGED')),
            cancelled_count=Count('id', filter=Q(status='CANCELLED')),
            total_cbm=Sum('cbm'),
            total_weight=Sum('weight')
        )

        # add ready/complete counts
        agg[f'{ready_status.lower()}_count'] = qs.filter(status=ready_status).count()
        agg[f'{complete_status.lower()}_count'] = qs.filter(status=complete_status).count()

        # set zeros for None
        for k, v in list(agg.items()):
            agg[k] = v or 0

        # average days_in_warehouse for active (non-completed) items
        non_completed = qs.exclude(status=complete_status)
        total_non_completed = non_completed.count()
        if total_non_completed:
            # small dataset compute exact average
            if total_non_completed <= 1000:
                total_days = sum((timezone.now().date() - item.date_received.date()).days for item in non_completed)
                avg_days = round(total_days / total_non_completed, 1)
            else:
                # approximate using oldest & newest
                oldest = non_completed.order_by('date_received').first()
                newest = non_completed.order_by('-date_received').first()
                if oldest and newest:
                    oldest_days = (timezone.now().date() - oldest.date_received.date()).days
                    newest_days = (timezone.now().date() - newest.date_received.date()).days
                    avg_days = round((oldest_days + newest_days) / 2, 1)
                else:
                    avg_days = 0.0
        else:
            avg_days = 0.0

        agg['average_days_in_warehouse'] = avg_days

        # processing rate: percent of items in ready+complete
        total_items = agg['total_count']
        ready_plus_complete = agg[f'{ready_status.lower()}_count'] + agg[f'{complete_status.lower()}_count']
        agg['processing_rate'] = round((ready_plus_complete / total_items) * 100, 2) if total_items else 0.0

        return Response(agg)

    @action(detail=False, methods=['get'])
    def flagged_items(self, request):
        qs = self.get_queryset().filter(status='FLAGGED')
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'count': qs.count(), 'items': serializer.data})
        serializer = self.get_serializer(qs, many=True)
        return Response({'count': qs.count(), 'items': serializer.data})

    @action(detail=False, methods=['get'])
    def ready_for_shipping(self, request):
        ready_status = 'READY_FOR_DELIVERY' if self.warehouse_type == 'ghana' else 'READY_FOR_SHIPPING'
        qs = self.get_queryset().filter(status=ready_status)
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'count': qs.count(), 'items': serializer.data})
        serializer = self.get_serializer(qs, many=True)
        return Response({'count': qs.count(), 'items': serializer.data})

    @action(detail=False, methods=['get'])
    def overdue_items(self, request):
        try:
            days = int(request.query_params.get('days', 30))
            if days < 0:
                raise ValueError
        except ValueError:
            return Response({'error': 'Invalid days parameter'}, status=status.HTTP_400_BAD_REQUEST)

        cutoff = timezone.now() - timedelta(days=days)
        if self.warehouse_type == 'ghana':
            active_statuses = ['PENDING', 'READY_FOR_DELIVERY', 'FLAGGED']
        else:
            active_statuses = ['PENDING', 'READY_FOR_SHIPPING', 'FLAGGED']

        qs = self.get_queryset().filter(date_received__lt=cutoff, status__in=active_statuses)
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response({'threshold_days': days, 'count': qs.count(), 'items': serializer.data})
        serializer = self.get_serializer(qs, many=True)
        return Response({'threshold_days': days, 'count': qs.count(), 'items': serializer.data})

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser], throttle_classes=[])
    def upload_excel(self, request):
        """
        Enhanced Excel upload to bulk create entries with specific column structure (A,B,C,D,E,G,H).
        Supports: shipping_mark, date_receipt, date_loading, description, quantity, cbm, supply_tracking.
        Uses atomic transaction for safety with detailed error reporting.
        """
        serializer = ExcelUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            processed = serializer.process_excel_file()
            if processed['warehouse_type'] != self.warehouse_type:
                return Response({'error': f'Use the {processed["warehouse_type"]} endpoint for this file.'},
                                status=status.HTTP_400_BAD_REQUEST)

            created = []
            failed = []
            errors = []
            created_details = []

            with transaction.atomic():
                for row_index, row in enumerate(processed['data'], 1):
                    shipping_mark = row.get('shipping_mark')
                    supply_tracking = row.get('supply_tracking')

                    try:
                        # Resolve customer by shipping_mark (optional for China uploads)
                        customer = None
                        if shipping_mark:
                            customer = CustomerUser.objects.filter(shipping_mark=shipping_mark).first()
                            if not customer and processed['warehouse_type'] != 'china':
                                errors.append(f"Row {row_index + 1}: Customer with shipping_mark '{shipping_mark}' not found")
                                failed.append(supply_tracking)
                                continue

                        # Prevent duplicate supply_tracking (unique constraint)
                        if self.model_class.objects.filter(supply_tracking=supply_tracking).exists():
                            errors.append(f"Row {row_index + 1}: Supply tracking '{supply_tracking}' already exists")
                            failed.append(supply_tracking)
                            continue

                        # Build kwargs to create object with new date fields
                        create_kwargs = {
                            'customer': customer,
                            'shipping_mark': (shipping_mark or 'UNKNOWN')[:MAX_SHIPPING_MARK_LENGTH],
                            'supply_tracking': supply_tracking,
                            'cbm': row.get('cbm'),
                            'weight': row.get('weight', None),
                            'quantity': row.get('quantity'),
                            'description': row.get('description', None),
                            'status': row.get('status', 'PENDING'),
                            'method_of_shipping': row.get('method_of_shipping', 'SEA'),
                            'date_loading': row.get('date_loading', None),
                        }
                        
                        # Handle date_received - if provided in Excel, override auto_now_add
                        if row.get('date_receipt'):
                            # Since date_received has auto_now_add=True, we need to handle this specially
                            # We'll store the Excel date in date_loading for now, or create a new field
                            pass  # Keep auto_now_add behavior for date_received
                        
                        # Ghana-specific fields
                        if hasattr(self.model_class, 'location'):
                            create_kwargs['location'] = row.get('location', 'ACCRA')

                        # Create the object
                        obj = self.model_class.objects.create(**create_kwargs)
                        
                        # Update date_received if provided in Excel (after creation)
                        if row.get('date_receipt'):
                            # Convert date to datetime for the datetime field
                            from django.utils import timezone
                            import datetime
                            date_receipt = row.get('date_receipt')
                            if isinstance(date_receipt, datetime.date):
                                date_receipt = timezone.make_aware(
                                    datetime.datetime.combine(date_receipt, datetime.time())
                                )
                            obj.date_received = date_receipt
                            obj.save(update_fields=['date_received'])
                        
                        created.append(obj.supply_tracking)
                        created_details.append({
                            'supply_tracking': obj.supply_tracking,
                            'shipping_mark': obj.shipping_mark,
                            'quantity': obj.quantity,
                            'status': obj.status
                        })
                        
                    except Exception as e:
                        logger.exception(f"Failed to create goods row for tracking {supply_tracking}")
                        errors.append(f"Row {row_index + 1}: Failed to create record for '{supply_tracking}': {str(e)}")
                        failed.append(supply_tracking)

            # Generate summary
            result = {
                'total_processed': processed['total_rows'],
                'successful_creates': len(created),
                'failed_creates': len(failed),
                'errors': errors,
                'created_items': created,
                'created_details': created_details,
                'validation_summary': {
                    'total_rows_in_file': processed['total_rows'],
                    'valid_rows': processed['valid_rows'],
                    'invalid_rows': processed['total_rows'] - processed['valid_rows'],
                    'duplicate_tracking_numbers': len([e for e in errors if 'already exists' in e]),
                    'missing_customers': len([e for e in errors if 'not found' in e]),
                    'other_errors': len([e for e in errors if 'already exists' not in e and 'not found' not in e])
                }
            }
            
            response_status = status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST
            response_message = f"Successfully processed {len(created)} out of {processed['total_rows']} rows"
            
            return Response({
                'success': len(created) > 0,
                'message': response_message,
                'results': BulkCreateResultSerializer(result).data
            }, status=response_status)

        except Exception as e:
            logger.exception("Excel processing error")
            return Response({
                'success': False,
                'error': str(e),
                'message': 'Failed to process Excel file'
            }, status=status.HTTP_400_BAD_REQUEST)
            logger.exception("Excel processing error")
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def get_template_data(self):
        raise NotImplementedError("Child classes must implement get_template_data()")

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download Excel template for bulk upload. Returns a professionally formatted template 
        with sample data that guarantees successful upload.
        """
        template_data = self.get_template_data()
        
        # Create DataFrame with proper column order (A,B,C,D,E,F,G,H)
        df = pd.DataFrame(template_data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the main data
            df.to_excel(writer, sheet_name=self.template_sheet_name, index=False, startrow=2)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets[self.template_sheet_name]
            
            # Add title and instructions
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Title
            worksheet['A1'] = f'{self.warehouse_type.title()} Warehouse - Goods Received Template'
            title_font = Font(size=16, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='1f4e79', end_color='1f4e79', fill_type='solid')
            worksheet['A1'].font = title_font
            worksheet['A1'].fill = title_fill
            worksheet['A1'].alignment = Alignment(horizontal='center')
            worksheet.merge_cells('A1:H1')
            
            # Instructions
            instructions = "IMPORTANT: Column position is critical! Follow A,B,C,D,E,G,H exactly. Delete sample data before adding yours."
            worksheet['A2'] = instructions
            worksheet['A2'].font = Font(size=10, italic=True, color='D32F2F')
            worksheet.merge_cells('A2:H2')
            
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Format headers (row 3)
            for col in range(1, 9):  # A through H
                cell = worksheet.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data formatting
            data_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Format data rows
            for row in range(4, 4 + len(template_data[list(template_data.keys())[0]])):
                for col in range(1, 9):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = data_border
                    if col in [2, 3]:  # Date columns
                        cell.alignment = Alignment(horizontal='center')
                    elif col == 5:  # Quantity column
                        cell.alignment = Alignment(horizontal='center')
                    elif col == 7:  # CBM column
                        cell.alignment = Alignment(horizontal='center')
            
            # Column widths
            column_widths = {
                'A': 18,  # Shipping Mark
                'B': 15,  # Date Receipt
                'C': 15,  # Date Loading
                'D': 45,  # Description
                'E': 12,  # Quantity
                'F': 15,  # Specifications
                'G': 12,  # CBM
                'H': 20   # Tracking No
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Add validation notes in a separate sheet
            notes_sheet = workbook.create_sheet('Upload Instructions')
            notes_data = [
                ['Column', 'Field Name', 'Required', 'Format/Rules', 'Example'],
                ['A', 'Shipping Mark/Client', 'YES', 'Customer identifier (max 20 chars)', 'PM001, PMJOHN01'],
                ['B', 'Date of Receipt', 'YES', 'DD/MM/YYYY format', '25/12/2024, 01/01/2025'],
                ['C', 'Date of Loading', 'NO', 'DD/MM/YYYY format or empty', '26/12/2024 or blank'],
                ['D', 'Description', 'YES', 'Text description of goods', 'Electronics components'],
                ['E', 'CTNS (Quantity)', 'YES', 'Number greater than 0', '5, 10, 25'],
                ['F', 'Specifications', 'NO', 'This column is IGNORED', 'Any text (ignored)'],
                ['G', 'CBM', 'YES*', 'Decimal number (*Required for Ghana)', '2.5, 10.0, 0.5'],
                ['H', 'Supplier Tracking No', 'YES', 'Unique tracking identifier', 'TRK123456789']
            ]
            
            for row_idx, row_data in enumerate(notes_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = notes_sheet.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # Header
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
            
            # Auto-fit columns in notes sheet
            for col in notes_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                notes_sheet.column_dimensions[column].width = adjusted_width
        
        output.seek(0)
        response = HttpResponse(output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{self.template_filename}"'
        return response

    # Light-weight analytics helpers (can be extended)
    def _calculate_accuracy_rate(self, queryset):
        total = queryset.count() or 1
        flagged = queryset.filter(status='FLAGGED').count()
        return round(((total - flagged) / total) * 100, 2)

    def _calculate_processing_speed(self, queryset):
        processed = queryset.filter(status__in=['SHIPPED', 'DELIVERED'])
        if not processed.exists():
            return 0.0
        total_days = sum((timezone.now().date() - p.date_received.date()).days for p in processed)
        return round(total_days / processed.count(), 2)

    @action(detail=False, methods=['get', 'post'])
    def sea_cargo(self, request):
        """Get goods filtered by SEA method_of_shipping or create new SEA cargo"""
        if request.method == 'GET':
            # Don't modify self.queryset, create a fresh filtered queryset
            queryset = self.get_queryset().filter(method_of_shipping='SEA')
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        
        elif request.method == 'POST':
            # Force method_of_shipping to SEA for this endpoint
            data = request.data.copy()
            data['method_of_shipping'] = 'SEA'
            
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=False, methods=['get', 'post'])
    def air_cargo(self, request):
        """Get goods filtered by AIR method_of_shipping or create new AIR cargo"""
        if request.method == 'GET':
            # Don't modify self.queryset, create a fresh filtered queryset
            queryset = self.get_queryset().filter(method_of_shipping='AIR')
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
            
        elif request.method == 'POST':
            # Force method_of_shipping to AIR for this endpoint
            data = request.data.copy()
            data['method_of_shipping'] = 'AIR'
            
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    @action(detail=False, methods=['get'])
    def sea_statistics(self, request):
        """Statistics for SEA cargo only"""
        qs = self.get_queryset().filter(method_of_shipping='SEA')
        # Use similar logic to statistics but with filtered queryset
        if self.warehouse_type == 'ghana':
            ready_status = 'READY_FOR_DELIVERY'
            complete_status = 'DELIVERED'
        else:
            ready_status = 'READY_FOR_SHIPPING'
            complete_status = 'SHIPPED'

        agg = qs.aggregate(
            total_count=Count('id'),
            pending_count=Count('id', filter=Q(status='PENDING')),
            flagged_count=Count('id', filter=Q(status='FLAGGED')),
            cancelled_count=Count('id', filter=Q(status='CANCELLED')),
            total_cbm=Sum('cbm'),
            total_weight=Sum('weight')
        )

        agg[f'{ready_status.lower()}_count'] = qs.filter(status=ready_status).count()
        agg[f'{complete_status.lower()}_count'] = qs.filter(status=complete_status).count()

        for k, v in list(agg.items()):
            agg[k] = v or 0

        return Response(agg)

    @action(detail=False, methods=['get'])
    def air_statistics(self, request):
        """Statistics for AIR cargo only"""
        qs = self.get_queryset().filter(method_of_shipping='AIR')
        # Use similar logic to statistics but with filtered queryset
        if self.warehouse_type == 'ghana':
            ready_status = 'READY_FOR_DELIVERY'
            complete_status = 'DELIVERED'
        else:
            ready_status = 'READY_FOR_SHIPPING'
            complete_status = 'SHIPPED'

        agg = qs.aggregate(
            total_count=Count('id'),
            pending_count=Count('id', filter=Q(status='PENDING')),
            flagged_count=Count('id', filter=Q(status='FLAGGED')),
            cancelled_count=Count('id', filter=Q(status='CANCELLED')),
            total_cbm=Sum('cbm'),
            total_weight=Sum('weight')
        )

        agg[f'{ready_status.lower()}_count'] = qs.filter(status=ready_status).count()
        agg[f'{complete_status.lower()}_count'] = qs.filter(status=complete_status).count()

        for k, v in list(agg.items()):
            agg[k] = v or 0

        return Response(agg)


class GoodsReceivedChinaViewSet(BaseGoodsReceivedViewSet):
    queryset = GoodsReceivedChina.objects.all()
    serializer_class = GoodsReceivedChinaSerializer
    model_class = GoodsReceivedChina
    warehouse_type = 'china'
    template_filename = 'china_goods_template.xlsx'
    template_sheet_name = 'China Goods Template'
    
    # Add method_of_shipping to filterset_fields to enable proper filtering
    filterset_fields = ['status', 'shipping_mark', 'supply_tracking', 'method_of_shipping']

    def get_template_data(self):
        current_date = timezone.now().strftime('%d/%m/%Y')
        yesterday = (timezone.now() - timedelta(days=1)).strftime('%d/%m/%Y')
        
        return {
            'SHIPPING MARK/CLIENT': ['PM001', 'PM002', 'PM003', 'PM004', 'PM005'],
            'DATE OF RECEIPT': [current_date, yesterday, current_date, yesterday, current_date],
            'DATE OF LOADING': ['', current_date, '', yesterday, ''],
            'DESCRIPTION': [
                'Electronics - Mobile phones and accessories',
                'Textiles - Cotton fabrics and clothing items',
                'Household - Kitchen appliances and cookware',
                'Automotive - Engine parts and components',
                'Furniture - Office chairs and tables'
            ],
            'CTNS': [5, 12, 8, 15, 3],
            'SPECIFICATIONS': ['N/A', 'N/A', 'N/A', 'N/A', 'N/A'],
            'CBM': [2.5, 4.8, 3.2, 6.1, 1.5],
            'SUPPLIER&TRACKING NO': ['TRK123456789', 'TRK987654321', 'TRK555666777', 'TRK888999000', 'TRK111222333']
        }


class GoodsReceivedGhanaViewSet(BaseGoodsReceivedViewSet):
    queryset = GoodsReceivedGhana.objects.all()
    serializer_class = GoodsReceivedGhanaSerializer
    model_class = GoodsReceivedGhana
    warehouse_type = 'ghana'
    template_filename = 'ghana_goods_template.xlsx'
    template_sheet_name = 'Ghana Goods Template'

    def get_template_data(self):
        current_date = timezone.now().strftime('%d/%m/%Y')
        yesterday = (timezone.now() - timedelta(days=1)).strftime('%d/%m/%Y')
        
        return {
            'SHIPPING MARK/CLIENT': ['PM001', 'PM002', 'PM003', 'PM004', 'PM005'],
            'DATE OF RECEIPT': [current_date, yesterday, current_date, yesterday, current_date],
            'DATE OF LOADING': ['', current_date, '', yesterday, ''],
            'DESCRIPTION': [
                'Textiles - Cotton fabrics and traditional clothing',
                'Food products - Canned goods and spices',
                'Cosmetics - Beauty products and skincare items',
                'Sports equipment - Soccer balls and athletic gear',
                'Books - Educational materials and stationery'
            ],
            'CTNS': [8, 20, 6, 12, 4],
            'SPECIFICATIONS': ['N/A', 'N/A', 'N/A', 'N/A', 'N/A'],
            'CBM': [3.2, 7.5, 2.1, 4.8, 1.8],
            'SUPPLIER&TRACKING NO': ['TRK555666777', 'TRK888999000', 'TRK111222333', 'TRK444555666', 'TRK777888999']
        }


# ---------------------------------------
# Customer-specific ReadOnly viewsets
# ---------------------------------------
class CustomerBaseGoodsReceivedViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only views for customers: only returns items for authenticated customer's shipping_mark.
    Child classes must set model_class and serializer_class.
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'location']
    search_fields = ['supply_tracking', 'description']
    ordering_fields = ['date_received', 'created_at', 'status']
    ordering = ['-date_received']

    model_class = None
    serializer_class = None

    def get_queryset(self):
        user = self.request.user
        return self.model_class.objects.filter(shipping_mark=user.shipping_mark)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.shipping_mark != request.user.shipping_mark:
            return Response({'detail': 'You can only access your own goods.'}, status=status.HTTP_403_FORBIDDEN)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_statistics(self, request):
        qs = self.get_queryset()
        stats = {
            'total_items': qs.count(),
            'by_status': list(qs.values('status').annotate(count=Count('id'))),
            'total_cbm': qs.aggregate(total=Sum('cbm'))['total'] or 0,
            'total_weight': qs.aggregate(total=Sum('weight'))['total'] or 0,
            'recent_items': qs.filter(date_received__gte=timezone.now() - timedelta(days=30)).count()
        }
        return Response(stats)

    @action(detail=False, methods=['get'])
    def my_flagged_items(self, request):
        qs = self.get_queryset().filter(status='FLAGGED')
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def ready_for_shipping(self, request):
        ready_status = 'READY_FOR_SHIPPING'
        qs = self.get_queryset().filter(status=ready_status)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def overdue_items(self, request):
        days = int(request.query_params.get('days', 30))
        cutoff = timezone.now() - timedelta(days=days)
        qs = self.get_queryset().filter(date_received__lt=cutoff, status__in=['PENDING', 'FLAGGED'])
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def tracking(self, request):
        tracking_id = request.query_params.get('tracking_id')
        if not tracking_id:
            return Response({'detail': 'tracking_id parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            item = self.get_queryset().get(supply_tracking=tracking_id)
            serializer = self.get_serializer(item)
            return Response(serializer.data)
        except self.model_class.DoesNotExist:
            return Response({'detail': 'Item not found or not accessible'}, status=status.HTTP_404_NOT_FOUND)


class CustomerGoodsReceivedChinaViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only view for customers to see ALL goods in China warehouse
    (not filtered by shipping mark - all customers can see all China goods)
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'method_of_shipping']  # China model doesn't have 'location'
    search_fields = ['supply_tracking', 'description', 'shipping_mark']
    ordering_fields = ['date_received', 'created_at', 'status']
    ordering = ['-date_received']
    serializer_class = GoodsReceivedChinaSerializer
    model_class = GoodsReceivedChina

    def get_queryset(self):
        # Return ALL China goods for any customer to see, limited to recent activity
        qs = self.model_class.objects.all()

        def parse_date(value):
            if not value:
                return None
            try:
                return datetime.strptime(value, "%Y-%m-%d").date()
            except (TypeError, ValueError):
                return None

        today = timezone.now().date()
        default_start = today - timedelta(days=30)

        requested_start = parse_date(self.request.query_params.get('date_from'))
        requested_end = parse_date(self.request.query_params.get('date_to'))

        start_date = requested_start or default_start
        qs = qs.filter(date_received__gte=start_date)

        if requested_end:
            qs = qs.filter(date_received__lte=requested_end)

        return qs

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Statistics for all China goods (not user-specific)"""
        qs = self.get_queryset()
        stats = {
            'total_items': qs.count(),
            'by_status': list(qs.values('status').annotate(count=Count('id'))),
            'total_cbm': qs.aggregate(total=Sum('cbm'))['total'] or 0,
            'total_weight': qs.aggregate(total=Sum('weight'))['total'] or 0,
            'recent_items': qs.filter(date_received__gte=timezone.now() - timedelta(days=30)).count()
        }
        return Response(stats)


class CustomerGoodsReceivedGhanaViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only view for customers to see ALL goods in Ghana warehouse
    (not filtered by shipping mark - all customers can see all Ghana goods)
    """
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'location']
    search_fields = ['supply_tracking', 'description', 'shipping_mark']
    ordering_fields = ['date_received', 'created_at', 'status']
    ordering = ['-date_received']
    serializer_class = GoodsReceivedGhanaSerializer
    model_class = GoodsReceivedGhana

    def get_queryset(self):
        # Return ALL Ghana goods for any customer to see
        return self.model_class.objects.all()

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Statistics for all Ghana goods (not user-specific)"""
        qs = self.get_queryset()
        stats = {
            'total_items': qs.count(),
            'by_status': list(qs.values('status').annotate(count=Count('id'))),
            'total_cbm': qs.aggregate(total=Sum('cbm'))['total'] or 0,
            'total_weight': qs.aggregate(total=Sum('weight'))['total'] or 0,
            'recent_items': qs.filter(date_received__gte=timezone.now() - timedelta(days=30)).count()
        }
        return Response(stats)


# ---------------------------------------
# Customer unified dashboard
# ---------------------------------------
from rest_framework.views import APIView


class CustomerDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        china_goods = GoodsReceivedChina.objects.filter(shipping_mark=user.shipping_mark)
        ghana_goods = GoodsReceivedGhana.objects.filter(shipping_mark=user.shipping_mark)

        china_stats = {
            'total_items': china_goods.count(),
            'total_cbm': china_goods.aggregate(total=Sum('cbm'))['total'] or 0,
            'total_weight': china_goods.aggregate(total=Sum('weight'))['total'] or 0,
            'by_status': list(china_goods.values('status').annotate(count=Count('id'))),
            'flagged_count': china_goods.filter(status='FLAGGED').count(),
            'ready_count': china_goods.filter(status='READY_FOR_SHIPPING').count(),
        }

        ghana_stats = {
            'total_items': ghana_goods.count(),
            'total_cbm': ghana_goods.aggregate(total=Sum('cbm'))['total'] or 0,
            'total_weight': ghana_goods.aggregate(total=Sum('weight'))['total'] or 0,
            'by_status': list(ghana_goods.values('status').annotate(count=Count('id'))),
            'flagged_count': ghana_goods.filter(status='FLAGGED').count(),
            'ready_count': ghana_goods.filter(status='READY_FOR_DELIVERY').count(),
        }

        recent_china = china_goods.order_by('-date_received')[:5]
        recent_ghana = ghana_goods.order_by('-date_received')[:5]

        threshold = timezone.now() - timedelta(days=30)
        overdue_china = china_goods.filter(date_received__lt=threshold, status__in=['PENDING', 'FLAGGED']).count()
        overdue_ghana = ghana_goods.filter(date_received__lt=threshold, status__in=['PENDING', 'FLAGGED']).count()

        dashboard = {
            'customer_info': {
                'shipping_mark': user.shipping_mark,
                'name': user.get_full_name(),
                'company': getattr(user, 'company_name', '')
            },
            'overview': {
                'total_items': china_stats['total_items'] + ghana_stats['total_items'],
                'total_cbm': china_stats['total_cbm'] + ghana_stats['total_cbm'],
                'total_weight': china_stats['total_weight'] + ghana_stats['total_weight'],
                'total_flagged': china_stats['flagged_count'] + ghana_stats['flagged_count'],
                'total_ready': china_stats['ready_count'] + ghana_stats['ready_count'],
                'total_overdue': overdue_china + overdue_ghana
            },
            'china_warehouse': china_stats,
            'ghana_warehouse': ghana_stats,
            'recent_items': {
                'china': GoodsReceivedChinaSerializer(recent_china, many=True).data,
                'ghana': GoodsReceivedGhanaSerializer(recent_ghana, many=True).data
            }
        }

        return Response(dashboard)


# ==========================================
# NEW CONTAINER-BASED VIEWS
# ==========================================

class GoodsReceivedContainerViewSet(viewsets.ModelViewSet):
    """ViewSet for managing goods received containers (separate from cargo containers)."""
    
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['container_type', 'location', 'status']
    search_fields = ['container_id', 'notes']
    ordering_fields = ['created_at', 'arrival_date', 'status']
    ordering = ['-created_at']
    
    def get_queryset(self):
        from .models import GoodsReceivedContainer
        from django.db.models import Q
        
        queryset = GoodsReceivedContainer.objects.all().prefetch_related('goods_items')
        
        # Custom search functionality - search in both container fields and item fields
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(container_id__icontains=search) |
                Q(notes__icontains=search) |
                Q(goods_items__shipping_mark__icontains=search) |
                Q(goods_items__supply_tracking__icontains=search) |
                Q(goods_items__description__icontains=search)
            ).distinct()
        
        return queryset
    
    def get_serializer_class(self):
        from .serializers import (
            GoodsReceivedContainerSerializer, 
            CreateGoodsReceivedContainerSerializer,
            UpdateGoodsReceivedContainerSerializer
        )
        if self.action == 'create':
            return CreateGoodsReceivedContainerSerializer
        elif self.action in ['update', 'partial_update']:
            return UpdateGoodsReceivedContainerSerializer
        return GoodsReceivedContainerSerializer
    
    def update(self, request, *args, **kwargs):
        """Override update to force partial=True for all updates."""
        kwargs['partial'] = True
        return super().update(request, *args, **kwargs)
    
    def partial_update(self, request, *args, **kwargs):
        """Override partial_update to ensure it uses the update serializer."""
        kwargs['partial'] = True
        return super().partial_update(request, *args, **kwargs)
    
    @action(detail=True, methods=['get'])
    def items(self, request, pk=None):
        """Get all items in a specific container, grouped by shipping mark."""
        container = self.get_object()
        from .serializers import GoodsReceivedItemSerializer
        
        items = container.goods_items.all().order_by('shipping_mark', 'created_at')
        
        # Group by shipping mark
        grouped_items = {}
        for item in items:
            mark = item.shipping_mark or "Unknown"
            if mark not in grouped_items:
                grouped_items[mark] = []
            grouped_items[mark].append(GoodsReceivedItemSerializer(item).data)
        
        return Response({
            'container': self.get_serializer(container).data,
            'items_by_shipping_mark': grouped_items,
            'total_items': items.count()
        })
    
    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
        """Add a new item to this container."""
        container = self.get_object()
        from .serializers import CreateGoodsReceivedItemSerializer
        
        data = request.data.copy()
        data['container'] = container.pk
        
        serializer = CreateGoodsReceivedItemSerializer(data=data)
        if serializer.is_valid():
            item = serializer.save()
            from .serializers import GoodsReceivedItemSerializer
            return Response(GoodsReceivedItemSerializer(item).data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get statistics for goods received containers."""
        from .models import GoodsReceivedContainer
        from django.db.models import Q, Count, Sum
        
        containers = GoodsReceivedContainer.objects.all()
        
        # Filter by location if specified
        location = request.query_params.get('location')
        if location:
            containers = containers.filter(location=location)
        
        # Filter by container type if specified
        container_type = request.query_params.get('container_type')
        if container_type:
            containers = containers.filter(container_type=container_type)
        
        stats = {
            'total_containers': containers.count(),
            'pending_containers': containers.filter(status='pending').count(),
            'processing_containers': containers.filter(status='processing').count(),
            'ready_for_delivery': containers.filter(status='ready_for_delivery').count(),
            'delivered_containers': containers.filter(status='delivered').count(),
            'flagged_containers': containers.filter(status='flagged').count(),
            
            'total_items': sum(c.total_items_count or 0 for c in containers),
            'total_weight': sum(c.total_weight or 0 for c in containers),
            'total_cbm': sum(c.total_cbm or 0 for c in containers),
            
            'air_containers': containers.filter(container_type='air').count(),
            'sea_containers': containers.filter(container_type='sea').count(),
        }
        
        from .serializers import GoodsReceivedContainerStatsSerializer
        return Response(GoodsReceivedContainerStatsSerializer(stats).data)


class GoodsReceivedItemViewSet(viewsets.ModelViewSet):
    """ViewSet for managing individual goods received items."""
    
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['container', 'shipping_mark', 'status', 'customer']
    search_fields = ['shipping_mark', 'supply_tracking', 'description', 'supplier_name']
    ordering_fields = ['created_at', 'date_received', 'status']
    ordering = ['-created_at']
    
    def get_queryset(self):
        from .models import GoodsReceivedItem
        return GoodsReceivedItem.objects.all().select_related('container', 'customer')
    
    def get_serializer_class(self):
        from .serializers import GoodsReceivedItemSerializer, CreateGoodsReceivedItemSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return CreateGoodsReceivedItemSerializer
        return GoodsReceivedItemSerializer
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """Update the status of a specific item."""
        item = self.get_object()
        new_status = request.data.get('status')
        
        if not new_status:
            return Response({'error': 'Status is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate status
        from .models import GoodsReceivedItem
        valid_statuses = [choice[0] for choice in GoodsReceivedItem.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return Response({'error': f'Invalid status. Valid options: {valid_statuses}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        item.status = new_status
        item.save()
        
        return Response(self.get_serializer(item).data)
    
    @action(detail=False, methods=['post'])
    def bulk_status_update(self, request):
        """Update status for multiple items."""
        item_ids = request.data.get('item_ids', [])
        new_status = request.data.get('status')
        
        if not item_ids or not new_status:
            return Response({'error': 'item_ids and status are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        from .models import GoodsReceivedItem
        valid_statuses = [choice[0] for choice in GoodsReceivedItem.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return Response({'error': f'Invalid status. Valid options: {valid_statuses}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        updated_count = GoodsReceivedItem.objects.filter(id__in=item_ids).update(status=new_status)
        
        return Response({
            'message': f'Updated {updated_count} items',
            'updated_count': updated_count
        })


# ==========================================
# LOCATION/TYPE SPECIFIC CONTAINER VIEWSETS
# ==========================================

class GhanaAirContainerViewSet(GoodsReceivedContainerViewSet):
    """Ghana Air containers specifically - filters containers by location=ghana and container_type=air"""
    
    def get_queryset(self):
        # Get the base queryset with search functionality from parent
        queryset = super().get_queryset()
        # Apply Ghana Air specific filters
        return queryset.filter(
            location='ghana', 
            container_type='air'
        )


class GhanaSeaContainerViewSet(GoodsReceivedContainerViewSet):
    """Ghana Sea containers specifically - filters containers by location=ghana and container_type=sea"""
    
    def get_queryset(self):
        # Get the base queryset with search functionality from parent
        queryset = super().get_queryset()
        # Apply Ghana Sea specific filters
        return queryset.filter(
            location='ghana', 
            container_type='sea'
        )


class ChinaAirContainerViewSet(GoodsReceivedContainerViewSet):
    """China Air containers specifically - filters containers by location=china and container_type=air"""
    
    def get_queryset(self):
        from .models import GoodsReceivedContainer
        return GoodsReceivedContainer.objects.filter(
            location='china', 
            container_type='air'
        ).prefetch_related('goods_items')


class ChinaSeaContainerViewSet(GoodsReceivedContainerViewSet):
    """China Sea containers specifically - filters containers by location=china and container_type=sea"""
    
    def get_queryset(self):
        from .models import GoodsReceivedContainer
        return GoodsReceivedContainer.objects.filter(
            location='china', 
            container_type='sea'
        ).prefetch_related('goods_items')
