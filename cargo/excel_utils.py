"""
Excel upload utilities for container items processing.
Handles shipping mark normalization, Excel parsing, and data validation.
"""
import re
import unicodedata
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation
import logging

logger = logging.getLogger(__name__)


def normalize_shipping_mark(raw_mark: str) -> str:
    """
    Normalize shipping mark value according to specifications:
    - Trim whitespace (including non-breaking spaces)
    - Convert to uppercase
    - Collapse multiple spaces to one
    - Remove leading/trailing punctuation (, . : ; / \\)
    - Remove invisible characters
    """
    if not raw_mark:
        return ""
    
    # Remove invisible characters and normalize unicode
    normalized = unicodedata.normalize('NFKD', raw_mark)
    normalized = ''.join(c for c in normalized if unicodedata.category(c) != 'Cf')
    
    # Trim all whitespace including non-breaking spaces
    normalized = normalized.strip()
    
    # Convert to uppercase
    normalized = normalized.upper()
    
    # Collapse multiple spaces to one
    normalized = re.sub(r'\s+', ' ', normalized)
    
    # Remove leading/trailing punctuation
    normalized = normalized.strip(',.:/;\\')
    
    return normalized


def split_multiple_shipping_marks(raw_mark: str) -> List[str]:
    """
    Split shipping mark if it contains multiple marks separated by /,;|
    Returns list of normalized shipping marks.
    """
    if not raw_mark:
        return []
    
    # Split by common delimiters
    marks = re.split(r'[/,;|]', raw_mark)
    
    # Normalize each mark and filter out empty ones
    normalized_marks = []
    for mark in marks:
        normalized = normalize_shipping_mark(mark)
        if normalized:
            normalized_marks.append(normalized)
    
    return normalized_marks if normalized_marks else [normalize_shipping_mark(raw_mark)]


def parse_quantity(value: Any) -> Optional[int]:
    """
    Parse quantity as integer. Accept "1", "1.0", "1,000" (commas as thousand separators).
    Returns None if cannot parse as positive integer.
    """
    if value is None:
        return None
    
    try:
        # Convert to string and clean
        str_value = str(value).strip()
        if not str_value:
            return None
        
        # Remove commas (thousand separators)
        str_value = str_value.replace(',', '')
        
        # Try to parse as float first (to handle "1.0" case)
        float_value = float(str_value)
        
        # Check if it's effectively an integer
        if float_value.is_integer() and float_value > 0:
            return int(float_value)
        
        return None
    except (ValueError, OverflowError):
        return None


def parse_cbm(value: Any) -> Optional[Decimal]:
    """
    Parse CBM as decimal. Accept either dot or comma decimal separators.
    Returns None if empty, raises exception if invalid.
    """
    if value is None or str(value).strip() == '':
        return None
    
    try:
        # Convert to string and clean
        str_value = str(value).strip()
        
        # Replace comma with dot for decimal separator
        str_value = str_value.replace(',', '.')
        
        # Parse as Decimal
        cbm_value = Decimal(str_value)
        
        # Must be positive
        if cbm_value < 0:
            raise InvalidOperation("CBM cannot be negative")
        
        return cbm_value
    except (InvalidOperation, ValueError):
        raise ValueError(f"Invalid CBM value: {value}")


def is_header_row(row_data: List[Any]) -> bool:
    """
    Detect if the row is a header row by checking for non-data text.
    Returns True if row contains header-like text.
    """
    if not row_data or len(row_data) < 5:
        return False
    
    # Common header texts (case insensitive)
    header_patterns = [
        'shipping mark', 'qty', 'quantity', 'description', 'cbm', 'tracking',
        'mark', 'desc', 'volume', 'number', 'tracking number'
    ]
    
    # Check if any cell contains header-like text
    for cell in row_data[:8]:  # Check first 8 columns
        if cell and isinstance(cell, str):
            cell_lower = cell.lower().strip()
            for pattern in header_patterns:
                if pattern in cell_lower:
                    return True
    
    # Check if quantity column (E) looks like header (not numeric)
    quantity_cell = row_data[4] if len(row_data) > 4 else None
    if quantity_cell and isinstance(quantity_cell, str):
        quantity_str = str(quantity_cell).strip().lower()
        if any(word in quantity_str for word in ['qty', 'quantity', 'amount', 'count']):
            return True
    
    return False


class ExcelRowParser:
    """Parses Excel rows according to the specified column mapping."""
    
    def __init__(self):
        self.invalid_rows = []
        self.processed_count = 0
    
    def parse_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse entire Excel file and return structured data.
        
        Returns:
        {
            'candidates': [list of item candidates],
            'invalid_rows': [list of invalid rows with reasons],
            'total_rows': int
        }
        """
        candidates = []
        
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active
            
            row_number = 0
            header_skipped = False
            
            for row in sheet.iter_rows(values_only=True):
                row_number += 1
                
                # Skip completely empty rows
                if not any(cell for cell in row):
                    continue
                
                # Auto-detect and skip header row
                if not header_skipped and is_header_row(list(row)):
                    header_skipped = True
                    continue
                
                # Parse this data row
                parsed_items = self.parse_row(list(row), row_number)
                candidates.extend(parsed_items)
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            raise ValueError(f"Failed to parse Excel file: {e}")
        
        return {
            'candidates': candidates,
            'invalid_rows': self.invalid_rows,
            'total_rows': row_number
        }
    
    def parse_row(self, row_data: List[Any], source_row_number: int) -> List[Dict[str, Any]]:
        """
        Parse a single row according to column mapping.
        Returns list of item candidates (multiple if shipping mark contains separators).
        """
        # Extract columns according to mapping
        # A(0)=shipping_mark, D(3)=description, E(4)=quantity, G(6)=cbm, H(7)=tracking
        
        if len(row_data) < 8:
            # Pad with None values if row is too short
            row_data.extend([None] * (8 - len(row_data)))
        
        raw_shipping_mark = row_data[0] if len(row_data) > 0 else None
        description = row_data[3] if len(row_data) > 3 else None
        quantity_raw = row_data[4] if len(row_data) > 4 else None
        cbm_raw = row_data[6] if len(row_data) > 6 else None
        tracking_number = row_data[7] if len(row_data) > 7 else None
        
        # Validate shipping mark
        if not raw_shipping_mark or not str(raw_shipping_mark).strip():
            self.invalid_rows.append({
                'source_row_number': source_row_number,
                'reason': 'missing_shipping_mark',
                'raw_data': {
                    'shipping_mark': raw_shipping_mark,
                    'description': description,
                    'quantity': quantity_raw,
                    'cbm': cbm_raw,
                    'tracking_number': tracking_number
                }
            })
            return []
        
        # Parse quantity
        quantity = parse_quantity(quantity_raw)
        if quantity is None:
            self.invalid_rows.append({
                'source_row_number': source_row_number,
                'reason': 'bad_quantity',
                'raw_data': {
                    'shipping_mark': raw_shipping_mark,
                    'description': description,
                    'quantity': quantity_raw,
                    'cbm': cbm_raw,
                    'tracking_number': tracking_number
                }
            })
            return []
        
        # Parse CBM
        try:
            cbm = parse_cbm(cbm_raw)
        except ValueError:
            self.invalid_rows.append({
                'source_row_number': source_row_number,
                'reason': 'bad_cbm',
                'raw_data': {
                    'shipping_mark': raw_shipping_mark,
                    'description': description,
                    'quantity': quantity_raw,
                    'cbm': cbm_raw,
                    'tracking_number': tracking_number
                }
            })
            return []
        
        # Split multiple shipping marks
        shipping_marks = split_multiple_shipping_marks(str(raw_shipping_mark))
        
        # Create candidates for each shipping mark
        candidates = []
        for shipping_mark in shipping_marks:
            candidate = {
                'source_row_number': source_row_number,
                'shipping_mark_normalized': shipping_mark,
                'original_shipping_mark_raw': str(raw_shipping_mark),
                'description': str(description).strip() if description else '',
                'quantity': quantity,
                'cbm': cbm,
                'tracking_number': str(tracking_number).strip() if tracking_number else ''
            }
            candidates.append(candidate)
        
        return candidates