"""
Customer Excel Upload Views

API endpoints for uploading and processing customer data from Excel files.
Handles file upload, validation, duplicate checking, and bulk customer creation.
"""

import tempfile
import os
import logging
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, serializers
from rest_framework.permissions import IsAuthenticated
from rest_framework.renderers import JSONRenderer
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from django.core.exceptions import ValidationError
from users.models import CustomerUser
from users.customer_excel_utils import (
    process_customer_excel_upload,
    create_customer_from_data,
)


logger = logging.getLogger(__name__)


class CustomerExcelUploadSerializer(serializers.Serializer):
    """Serializer for customer Excel file upload."""
    
    file = serializers.FileField(
        required=True,
        help_text="Excel file containing customer data"
    )
    
    def validate_file(self, value):
        """Validate uploaded file."""
        # Check file extension
        if not value.name.lower().endswith(('.xlsx', '.xls')):
            raise serializers.ValidationError(
                "File must be an Excel file (.xlsx or .xls)"
            )
        
        # Check file size using centralized config (max 50MB)
        from excel_config import validate_file_size
        is_valid, error_msg = validate_file_size(value.size)
        if not is_valid:
            raise serializers.ValidationError(error_msg)
        
        return value


class CustomerBulkCreateSerializer(serializers.Serializer):
    """Serializer for bulk creating customers from validated data."""
    
    customers = serializers.ListField(
        child=serializers.DictField(),
        required=True,
        help_text="List of validated customer data to create"
    )
    
    def validate_customers(self, value):
        """Validate customer data list."""
        if not value:
            raise serializers.ValidationError("Customer list cannot be empty")
        
        # Increased limit to handle 4,000-7,000 customer uploads
        if len(value) > 10000:
            raise serializers.ValidationError("Cannot create more than 10,000 customers at once. Please split large files.")
        
        # Validate each customer data structure
        required_fields = ['shipping_mark', 'first_name', 'last_name', 'phone_normalized', 'source_row_number']
        
        for i, customer in enumerate(value):
            if not isinstance(customer, dict):
                raise serializers.ValidationError(f"Customer {i+1} must be a dictionary")
            
            for field in required_fields:
                if field not in customer or not customer[field]:
                    raise serializers.ValidationError(f"Customer {i+1} missing required field: {field}")
        
        return value


class CustomerExcelUploadView(APIView):
    """
    API endpoint for uploading and processing customer Excel files.
    """
    permission_classes = [IsAuthenticated]
    renderer_classes = [JSONRenderer]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        """
        Upload and process Excel file containing customer data with MEMORY-OPTIMIZED chunked processing.
        
        OPTIMIZATION FOR RENDER FREE TIER (512MB RAM):
        - Processes Excel file in chunks to minimize memory usage
        - Validates data in chunks without loading entire file
        - Memory-safe for files with 4,000-7,000 customer records
        
        Expected Excel format:
        - Column A: Shipping Mark (required, unique)
        - Column B: First Name (required)
        - Column C: Last Name (required)
        - Column D: Email (optional)
        - Column E: Phone Number (required, unique)
        
        Returns processing results including valid candidates and duplicates.
        """
        logger.info(f"[EXCEL-UPLOAD-START] User: {getattr(request.user, 'id', 'unknown')}")
        
        # Validate input
        serializer = CustomerExcelUploadSerializer(data=request.data)
        if not serializer.is_valid():
            logger.debug("Customer Excel upload validation failed", extra={
                "errors": serializer.errors,
                "user_id": getattr(request.user, "id", None),
            })
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        uploaded_file = serializer.validated_data['file']
        file_size_mb = uploaded_file.size / (1024 * 1024)
        
        logger.info(f"[EXCEL-UPLOAD-FILE] Name: {uploaded_file.name}, Size: {file_size_mb:.2f}MB")
        
        # Save uploaded file temporarily
        temp_file_path = None
        try:
            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                for chunk in uploaded_file.chunks():
                    temp_file.write(chunk)
                temp_file_path = temp_file.name
            
            logger.info(f"[EXCEL-UPLOAD-PROCESS] Starting chunked processing: {temp_file_path}")

            # Process the Excel file with memory optimization
            results = process_customer_excel_upload(temp_file_path)
            
            logger.info(f"[EXCEL-UPLOAD-PARSED] Success: {results.get('success')}, "
                       f"Valid: {results.get('parsing_results', {}).get('valid_candidates', 0)}")
            
            if not results['success']:
                logger.warning(
                    "Customer Excel upload returned unsuccessful parse",
                    extra={
                        "user_id": getattr(request.user, "id", None),
                        "message": results.get('message'),
                        "parsing_summary": results.get('parsing_results', {}),
                    }
                )
                return Response({
                    'success': False,
                    'message': results['message'],
                    'parsing_results': results['parsing_results']
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Return processing results
            response_data = {
                'success': True,
                'message': results['message'],
                'parsing_results': results['parsing_results'],
                'duplicate_results': results['duplicate_results'],
                'upload_id': f"customer_upload_{request.user.id}_{temp_file.name.split('/')[-1]}"
            }
            
            logger.info(f"[EXCEL-UPLOAD-COMPLETE] Valid candidates: {results['parsing_results']['valid_candidates']}")
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.exception(
                "[EXCEL-UPLOAD-ERROR] Processing failed",
                extra={
                    "user_id": getattr(request.user, "id", None),
                    "file_name": getattr(uploaded_file, "name", "unknown"),
                    "error": str(e),
                }
            )
            return Response({
                'success': False,
                'message': f'File processing failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        finally:
            # Clean up temporary file
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                    logger.debug(f"[EXCEL-UPLOAD-CLEANUP] Temp file deleted: {temp_file_path}")
                except Exception as cleanup_error:
                    logger.warning(f"[EXCEL-UPLOAD-CLEANUP] Failed to delete temp file: {cleanup_error}")


class CustomerBulkCreateView(APIView):
    """
    API endpoint for bulk creating customers from validated Excel data.
    """
    permission_classes = [IsAuthenticated]
    renderer_classes = [JSONRenderer]
    
    def post(self, request):
        """
        Create multiple customers from validated Excel data.
        
        Expects a list of validated customer data objects.
        """
        # CRITICAL: Early crash protection - catch ANY exception before it crashes the worker
        try:
            logger.info(f"Customer bulk create request started from user: {getattr(request.user, 'id', 'unknown')}")
            print(f"Bulk create request received. Method: {request.method}")
            
            # Check if request.data exists and is accessible
            if not hasattr(request, 'data'):
                logger.error("Request has no data attribute")
                return Response({
                    'success': False,
                    'message': 'Invalid request format',
                    'created_customers': [],
                    'failed_customers': []
                }, status=status.HTTP_400_BAD_REQUEST)
            
            print(f"Request data keys: {list(request.data.keys()) if request.data else 'Empty data'}")
            print(f"Request content type: {request.content_type}")
            
            # Validate input
            serializer = CustomerBulkCreateSerializer(data=request.data)
            if not serializer.is_valid():
                print(f"Serializer validation failed: {serializer.errors}")
                return Response({
                    'success': False,
                    'errors': serializer.errors,
                    'message': 'Invalid request data'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            customers_data = serializer.validated_data['customers']
            print(f"Number of customers to create: {len(customers_data)}")
            
            # Prevent memory issues on Render free tier
            # Increased limit to handle 4,000-7,000 customer uploads
            if len(customers_data) > 10000:
                return Response({
                    'success': False,
                    'error': f'Too many customers ({len(customers_data)}). Maximum 10,000 customers per request.',
                    'message': 'Request too large - please split into smaller files'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if len(customers_data) == 0:
                return Response({
                    'success': False,
                    'message': 'No customers provided',
                    'created_customers': [],
                    'failed_customers': []
                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception("Error validating bulk create request")
            return Response({
                'success': False,
                'message': f'Request validation failed: {str(e)}',
                'created_customers': [],
                'failed_customers': []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Main processing block with comprehensive error handling
        created_customers = []
        failed_customers = []
        
        try:
            logger.info(f"[BULK-CREATE-START] Processing {len(customers_data)} customers")
            print(f"Starting bulk creation of {len(customers_data)} customers")
            
            # ═══════════════════════════════════════════════════════════════
            # MEMORY-OPTIMIZED BATCH PROCESSING FOR RENDER FREE TIER
            # ═══════════════════════════════════════════════════════════════
            # 
            # STRATEGY:
            # - Process in small batches (25 customers) to minimize memory spikes
            # - Individual transactions prevent deadlocks on Postgres
            # - Progress logging helps track long-running operations
            # - Continue on errors (graceful degradation)
            # 
            # MEMORY PROFILE:
            # - Each customer object: ~2-5KB
            # - Batch of 25: ~125KB
            # - Total for 7,000: ~35MB in memory (spread across batches)
            # 
            batch_size = 25  # Optimized for 512MB RAM free tier
            total_customers = len(customers_data)
            total_batches = (total_customers + batch_size - 1) // batch_size
            
            logger.info(f"[BULK-CREATE-BATCHES] Total: {total_batches} batches of {batch_size}")
            
            for batch_start in range(0, total_customers, batch_size):
                batch_end = min(batch_start + batch_size, total_customers)
                batch = customers_data[batch_start:batch_end]
                batch_num = (batch_start // batch_size) + 1
                
                # Log progress every 20 batches (every 500 customers)
                if batch_num % 20 == 0 or batch_num == 1:
                    logger.info(f"[BULK-CREATE-PROGRESS] Batch {batch_num}/{total_batches} | "
                               f"Processed: {batch_start}/{total_customers} | "
                               f"Created: {len(created_customers)} | Failed: {len(failed_customers)}")
                
                print(f"Processing batch {batch_num}/{total_batches}: customers {batch_start + 1} to {batch_end}")
                
                for customer_data in batch:
                    try:
                        # Validate customer data structure
                        if not isinstance(customer_data, dict):
                            failed_customers.append({
                                'customer_data': customer_data,
                                'error': 'Invalid customer data format',
                                'source_row_number': customer_data.get('source_row_number', 0)
                            })
                            continue
                        
                        # Create customer with individual transaction
                        # This prevents nested transaction deadlocks on Render
                        with transaction.atomic():
                            customer = self._create_customer(customer_data, request.user)
                            created_customers.append({
                                'id': customer.id,
                                'shipping_mark': customer.shipping_mark,
                                'name': customer.get_full_name(),
                                'email': customer.email or '',
                                'phone': customer.phone,
                                'source_row_number': customer_data.get('source_row_number', 0)
                            })
                        
                    except Exception as e:
                        # Log the error but continue processing
                        logger.error(f"[BULK-CREATE-FAIL] Row {customer_data.get('source_row_number', '?')}: {str(e)}")
                        failed_customers.append({
                            'customer_data': customer_data,
                            'error': str(e),
                            'source_row_number': customer_data.get('source_row_number', 0)
                        })
                
                # Log batch completion (every batch for visibility)
                if batch_num % 20 == 0 or batch_num == total_batches:
                    print(f"Batch {batch_num}/{total_batches} completed. Total created: {len(created_customers)}, Total failed: {len(failed_customers)}")            # Return results
            response_data = {
                'success': True,
                'message': f'Processing completed. Created {len(created_customers)} customers, {len(failed_customers)} failed',
                'created_customers': created_customers,
                'failed_customers': failed_customers,
                'total_attempted': len(customers_data),
                'total_created': len(created_customers),
                'total_failed': len(failed_customers)
            }
        
            # If no customers were created but we had data, this suggests a systematic issue
            if len(customers_data) > 0 and len(created_customers) == 0:
                response_data['success'] = False
                response_data['message'] = 'Failed to create any customers - check error details'
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"Customer bulk create completed: {len(created_customers)} created, {len(failed_customers)} failed")
            return Response(response_data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            # FINAL SAFETY NET: Catch any uncaught exception in processing
            logger.exception(f"CRITICAL: Uncaught exception in customer bulk create: {str(e)}")
            return Response({
                'success': False,
                'message': f'Server error during processing: {str(e)}',
                'created_customers': created_customers,
                'failed_customers': failed_customers,
                'error_type': type(e).__name__
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _create_customer(self, customer_data, created_by_user):
        """
        Create a single customer from validated data.
        
        Args:
            customer_data: Dict containing customer information
            created_by_user: User who initiated the bulk creation
            
        Returns:
            Created CustomerUser instance
            
        Raises:
            ValueError: If customer creation fails
        """
        try:
            # Prepare payload for shared utility
            payload = {
                'shipping_mark': customer_data.get('shipping_mark'),
                'shipping_mark_normalized': customer_data.get('shipping_mark_normalized'),
                'first_name': customer_data.get('first_name'),
                'last_name': customer_data.get('last_name'),
                'phone_normalized': customer_data.get('phone_normalized'),
                'email': customer_data.get('email'),
                'region': customer_data.get('region') or 'GREATER_ACCRA',
                'accessible_warehouses': customer_data.get('accessible_warehouses', []),
            }

            # Create the customer using shared utility
            customer = create_customer_from_data(payload, created_by_user)
            print(f"Successfully created customer: {customer.shipping_mark} - {customer.get_full_name()}")
            return customer
            
        except ValidationError as e:
            # Django validation errors
            error_msg = f"Validation error: {str(e)}"
            logger.error(error_msg, extra={
                "customer_data": customer_data,
                "created_by": getattr(created_by_user, "id", None),
            })
            raise ValueError(error_msg)
            
        except Exception as e:
            # Any other errors - log and re-raise
            error_msg = f"Customer creation failed: {str(e)}"
            logger.exception(error_msg, extra={
                "customer_data": customer_data,
                "created_by": getattr(created_by_user, "id", None),
            })
            raise ValueError(error_msg)


class CustomerExcelTemplateView(APIView):
    """
    API endpoint for downloading customer Excel upload template.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """
        Generate and return a sample Excel template for customer upload.
        """
        try:
            import openpyxl
            from django.http import HttpResponse
            
            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Customer Upload Template"
            
            # Add headers
            headers = [
                'Shipping Mark',
                'First Name', 
                'Last Name',
                'Email',
                'Phone Number'
            ]
            
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            # Add sample data
            sample_data = [
                ['PM4 SAMPLE', 'John', 'Doe', 'john.doe@example.com', '+233241234567'],
                ['PM5 TEST', 'Jane', 'Smith', 'jane.smith@example.com', '+233551234567'],
                ['PM6 DEMO', 'Bob', 'Johnson', '', '+233201234567']
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Add instructions sheet
            instructions_sheet = workbook.create_sheet("Instructions")
            instructions = [
                "Customer Excel Upload Instructions:",
                "",
                "Column A: Shipping Mark (required, unique)",
                "- Alphanumeric identifier for the customer",
                "- Must be unique across all customers",
                "- 2-20 characters long",
                "",
                "Column B: First Name (required)",
                "- Customer's first name",
                "",
                "Column C: Last Name (required)", 
                "- Customer's last name",
                "",
                "Column D: Email (optional)",
                "- Valid email address",
                "- Must be unique if provided",
                "",
                "Column E: Phone Number (required, unique)",
                "- Customer's phone number",
                "- Include country code (e.g., +233)",
                "- Must be unique across all customers",
                "",
                "Notes:",
                "- First row should contain headers (will be skipped)",
                "- Empty rows will be ignored",
                "- Maximum 1000 customers per upload",
                "- Duplicates will be reported and skipped"
            ]
            
            for row_idx, instruction in enumerate(instructions, 1):
                instructions_sheet.cell(row=row_idx, column=1, value=instruction)
            
            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="customer_upload_template.xlsx"'
            
            workbook.save(response)
            return response
            
        except Exception as e:
            return Response({
                'success': False,
                'message': f'Template generation failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CustomerTestCreateView(APIView):
    """
    Simple test endpoint to verify customer creation works.
    """
    permission_classes = [IsAuthenticated]
    renderer_classes = [JSONRenderer]
    
    def post(self, request):
        """Test creating a single customer with minimal data."""
        try:
            import time
            timestamp = int(time.time())
            
            # Create a test customer
            test_customer = CustomerUser.objects.create_user(
                phone=f'+233{timestamp % 1000000000}',  # Unique phone
                first_name='Test',
                last_name='Customer',
                shipping_mark=f'TEST{timestamp % 10000}',  # Unique shipping mark
                user_role='CUSTOMER',
                user_type='INDIVIDUAL',
                region='GREATER_ACCRA',
                is_active=True,
                is_verified=False,
                created_by=request.user
            )
            
            return Response({
                'success': True,
                'message': f'Test customer created successfully',
                'customer': {
                    'id': test_customer.id,
                    'shipping_mark': test_customer.shipping_mark,
                    'name': test_customer.get_full_name(),
                    'phone': test_customer.phone
                }
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            import traceback
            return Response({
                'success': False,
                'message': f'Test customer creation failed: {str(e)}',
                'traceback': traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CustomerUploadStatsView(APIView):
    """
    API endpoint for getting customer upload statistics.
    """
    permission_classes = [IsAuthenticated]
    renderer_classes = [JSONRenderer]
    
    def get(self, request):
        """
        Get statistics about customer uploads and current customer data.
        """
        try:
            from django.db.models import Count, Q
            from django.utils import timezone
            from datetime import timedelta
            
            # Basic customer stats
            total_customers = CustomerUser.objects.filter(user_role='CUSTOMER').count()
            active_customers = CustomerUser.objects.filter(
                user_role='CUSTOMER', 
                is_active=True
            ).count()
            verified_customers = CustomerUser.objects.filter(
                user_role='CUSTOMER',
                is_verified=True
            ).count()
            
            # Recent additions (last 30 days)
            thirty_days_ago = timezone.now() - timedelta(days=30)
            recent_customers = CustomerUser.objects.filter(
                user_role='CUSTOMER',
                date_joined__gte=thirty_days_ago
            ).count()
            
            # Customers by region
            region_stats = CustomerUser.objects.filter(
                user_role='CUSTOMER'
            ).values('region').annotate(
                count=Count('id')
            ).order_by('-count')
            
            # Upload history (created by admins)
            upload_history = CustomerUser.objects.filter(
                user_role='CUSTOMER',
                created_by__isnull=False
            ).values(
                'created_by__first_name',
                'created_by__last_name'
            ).annotate(
                count=Count('id')
            ).order_by('-count')[:10]
            
            response_data = {
                'total_customers': total_customers,
                'active_customers': active_customers,
                'verified_customers': verified_customers,
                'recent_customers': recent_customers,
                'region_distribution': list(region_stats),
                'upload_history': list(upload_history)
            }
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'message': f'Stats retrieval failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ═══════════════════════════════════════════════════════════════
# ASYNC BACKGROUND TASK ENDPOINTS (RENDER TIMEOUT FIX)
# ═══════════════════════════════════════════════════════════════

class CustomerBulkCreateAsyncView(APIView):
    """
    Async bulk customer creation endpoint.
    
    SOLVES: Render's 60-100 second timeout limit for large uploads.
    
    Flow:
    1. Receives customer data from frontend
    2. Queues background task immediately
    3. Returns task ID in < 5 seconds
    4. Frontend polls /api/auth/customers/bulk-create/status/<task_id>/
    5. Background worker processes customers asynchronously
    
    Supports: 4,000-7,000 customer uploads without timeout
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """Queue bulk customer creation as background task."""
        try:
            from django_q.tasks import async_task
            import uuid
            
            logger.info(f"[ASYNC-BULK-CREATE-REQUEST] User: {request.user.id}")
            
            # Validate request
            customers_data = request.data.get('customers', [])
            if not customers_data or not isinstance(customers_data, list):
                return Response({
                    'success': False,
                    'message': 'Invalid request: "customers" array required'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Generate unique task ID
            task_id = str(uuid.uuid4())
            
            logger.info(
                f"[ASYNC-BULK-CREATE-QUEUE] Task: {task_id} | "
                f"Customers: {len(customers_data)} | User: {request.user.id}"
            )
            
            # Queue the background task
            from .async_customer_tasks import bulk_create_customers_task
            async_task(
                bulk_create_customers_task,
                customers_data,
                request.user.id,
                task_id,
                task_name=f'bulk_create_{task_id}',
                group='customer_bulk_create'
            )
            
            return Response({
                'success': True,
                'message': f'Bulk creation queued: {len(customers_data)} customers',
                'task_id': task_id,
                'total_customers': len(customers_data),
                'estimated_time_seconds': len(customers_data) // 10  # ~10 customers/sec
            }, status=status.HTTP_202_ACCEPTED)
            
        except Exception as e:
            logger.error(f"[ASYNC-BULK-CREATE-ERROR] {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'message': f'Failed to queue task: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CustomerBulkCreateStatusView(APIView):
    """
    Check status of async bulk customer creation task.
    
    Frontend polls this endpoint every 2-3 seconds to track progress.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, task_id):
        """Get task status and results."""
        try:
            from django_q.models import Task
            
            # Find task in database
            try:
                task = Task.objects.get(name=f'bulk_create_{task_id}')
            except Task.DoesNotExist:
                return Response({
                    'success': False,
                    'message': 'Task not found',
                    'status': 'NOT_FOUND'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Check task status
            if task.success is None:
                # Task still running
                return Response({
                    'success': True,
                    'status': 'RUNNING',
                    'message': 'Task is processing...',
                    'task_id': task_id
                }, status=status.HTTP_200_OK)
            
            elif task.success:
                # Task completed successfully
                result = task.result
                return Response({
                    'success': True,
                    'status': 'COMPLETE',
                    'message': 'Bulk creation complete',
                    'task_id': task_id,
                    'created': result.get('created', 0),
                    'failed': result.get('failed', 0),
                    'errors': result.get('errors', [])
                }, status=status.HTTP_200_OK)
            
            else:
                # Task failed
                return Response({
                    'success': False,
                    'status': 'FAILED',
                    'message': 'Task failed',
                    'task_id': task_id,
                    'error': str(task.result)
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Exception as e:
            logger.error(f"[ASYNC-STATUS-ERROR] Task: {task_id} | {str(e)}", exc_info=True)
            return Response({
                'success': False,
                'message': f'Status check failed: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)